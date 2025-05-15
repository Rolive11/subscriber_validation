# VS.py - Version 1.1.0
import pandas as pd
import os
import shutil
import sys
import re
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime

# Configuration from validate_subscribers.py
VALID_STATES = ["AL", "AK", "AZ", "AR", "CA", "CO", "CT", "DE", "FL", "GA", "HI", "ID", "IL", "IN", "IA", "KS", "KY", "LA", "ME", "MD", "MA", "MI", "MN", "MS", "MO", "MT", "NE", "NV", "NH", "NJ", "NM", "NY", "NC", "ND", "OH", "OK", "OR", "PA", "RI", "SC", "SD", "TN", "TX", "UT", "VT", "VA", "WA", "WV", "WI", "WY", "DC", "PR", "VI", "GU", "AS", "MP"]
VALID_TECHNOLOGIES = ["fiber", "cable", "dsl", "wireless_licensed", "wireless_unlicensed", "copper"]
EXPECTED_COLUMNS = ["customer", "lat", "lon", "address", "city", "state", "zip", "download", "upload", "voip_lines_quantity", "business_customer", "technology"]

# Street endings and patterns
MULTI_WORD_ENDINGS = (
    r"\bUS Highway\b|\bUS Hwy\b|\bPrivate Road\b|\bCounty Road\b|\bCounty Rd\b|\bCo Rd\b|\bState Route\b|"
    r"\bFarm to Market\b|\bCounty Hwy \d+\b|\bCounty FM \d+\b|\bFM Road \d+\b|"
    r"\bFire District \d+ Rd\b|\bState Hwy \d+\b|\bKamehameha Hwy\b|\bMamalahoa Hwy\b|"
    r"\bRoute C-\d+\b|\bRoute [A-Z]{2}\b|\b[A-Z]{2} Road\b|\bRS \d+\b|\bKY RS \d+\b"
)
SINGLE_WORD_ENDINGS = (
    r"\bAlley\b|\bALY\b|\bAvenue\b|\bAve\b|\bAv\b|\bBoulevard\b|\bBlvd\b|\bCircle\b|\bCir\b|\bCr\b|"
    r"\bCourt\b|\bCt\b|\bDrive\b|\bDr\b|\bExpressway\b|\bExpy\b|\bFM\b|\bHighway\b|\bHwy\b|"
    r"\bLane\b|\bLn\b|\bLoop\b|\bParkway\b|\bPkwy\b|\bPlace\b|\bPl\b|\bRoad\b|\bRd\b|\bRoute\b|\bRte\b|\bRt\b|\bSquare\b|"
    r"\bSq\b|\bStreet\b|\bSt\b|\bTerrace\b|\bTer\b|\bTrail\b|\bTrl\b|\bTurnpike\b|\bTpke\b|\bWay\b|\bWy\b|"
    r"\bCR\b|\bSR\b|\bFM\b|\bUS\b|\bInterstate\b|\bI-\b|"
    r"\bAZ-\d+\b|\bCA-\d+\b|\bCT-\d+\b|\bDE-\d+\b|\bFL-\d+\b|\bGA-\d+\b|\bID-\d+\b|"
    r"\bIL-\d+\b|\bIN-\d+\b|\bK-\d+\b|\bME-\d+\b|\bMD-\d+\b|\bMA-\d+\b|\bM-\d+\b|"
    r"\bMN-\d+\b|\bMS-\d+\b|\bNH-\d+\b|\bNJ-\d+\b|\bNM-\d+\b|\bNY-\d+\b|\bNC-\d+\b|"
    r"\bOH-\d+\b|\bOK-\d+\b|\bOR-\d+\b|\bPA-\d+\b|\bRI-\d+\b|\bSC-\d+\b|\bTN-\d+\b|"
    r"\bUT-\d+\b|\bVT-\d+\b|\bVA-\d+\b|\bWA-\d+\b|\bWV-\d+\b|\bWI-\d+\b|\bWY-\d+\b|"
    r"\bSH-\d+\b|\bC-\d+\b|\bCarr \d+\b|\bRoute \d+\b|\bCH \d+\b"
)
SPECIFIC_ROAD_PATTERN = r"(?i)(?:\d+\s+)?(?:County\s*(?:Road|Rd|CR)|Private\s*Road|Us\s*Hwy|Farm\s*to\s*Market|Farm\s*Road|Farm\s*to\s*Market\s*Road|FM\s*Rd|State\s*(?:Road|Rd|Route)|Old\s*State\s*(?:Road|Rd)|" \
                        r"(?:AL|AK|AZ|AR|CA|CO|CT|DE|FL|GA|HI|ID|IL|IN|IA|KS|KY|LA|ME|MD|MA|MI|MN|MS|MO|MT|NE|NV|NH|NJ|NM|NY|NC|ND|OH|OK|OR|PA|RI|SC|SD|TN|TX|UT|VT|VA|WA|WV|WI|WY|DC|PR|VI|GU|AS|MP)-\d+|" \
                        r"(?:Alabama|Alaska|Arizona|Arkansas|California|Colorado|Connecticut|Delaware|Florida|Georgia|Hawaii|Idaho|Illinois|Indiana|Iowa|Kansas|Kentucky|Louisiana|Maine|Maryland|Massachusetts|Michigan|Minnesota|Mississippi|Missouri|Montana|Nebraska|Nevada|New\sHampshire|New\sJersey|New\sMexico|New\sYork|North\sCarolina|North\sDakota|Ohio|Oklahoma|Oregon|Pennsylvania|Rhode\sIsland|South\sCarolina|South\sDakota|Tennessee|Texas|Utah|Vermont|Virginia|Washington|West\sVirginia|Wisconsin|Wyoming|District\sof\sColumbia|Puerto\sRico|Virgin\sIslands|Guam|American\sSamoa|Northern\sMariana\sIslands)\s*(?:Hwy|Highway|Route|Rte|Rt)\s*\d+)\s*(?:\d+(?:\s*(?:North|South|East|West|Northeast|Northwest|Southeast|Southwest|N|S|E|W|NE|NW|SE|SW))?)?\b"
STREET_ENDINGS = f"({MULTI_WORD_ENDINGS})|({SINGLE_WORD_ENDINGS})"
PO_BOX = r"\bPO Box\b|\bP\.O\. Box\b|\bPost Office Box\b"
RURAL_ROUTES = r"\bRR \d+ Box \d+\b|\bRural Route \d+ Box \d+\b|\bR\.R\. \d+ Box \d+\b|\bHC \d+ Box \d+\b"
FORBIDDEN_CHARS = r'[!@#$%^&*()+={}[\]|\"\'?/:;<,>]'

STATE_LON_RANGES = {
    "AL": (-88.473227, -84.889080), "AK": (-179.148909, 179.778470), "AZ": (-114.816510, -109.045223),
    "AR": (-94.617919, -89.644395), "CA": (-124.409591, -114.131211), "CO": (-109.060253, -102.041524),
    "CT": (-73.727775, -71.786994), "DE": (-75.788658, -75.048939), "FL": (-87.634896, -80.031056),
    "GA": (-85.605165, -80.840141), "HI": (-178.334698, -154.806773), "ID": (-117.243027, -111.043564),
    "IL": (-91.513079, -87.494756), "IN": (-88.097892, -84.787981), "IA": (-96.639704, -90.140061),
    "KS": (-102.051744, -94.588413), "KY": (-89.571510, -81.964971), "LA": (-94.043147, -88.817017),
    "ME": (-71.083924, -66.949895), "MD": (-79.487651, -75.048939), "MA": (-73.508142, -69.928393),
    "MI": (-90.418136, -82.413474), "MN": (-97.239209, -89.483385), "MS": (-91.655009, -88.097892),
    "MO": (-95.774704, -89.098843), "MT": (-116.050002, -104.039138), "NE": (-104.053514, -95.308290),
    "NV": (-120.005746, -114.039648), "NH": (-72.557247, -70.610621), "NJ": (-75.559614, -73.893979),
    "NM": (-109.050173, -103.001964), "NY": (-79.762152, -71.856214), "NC": (-84.321869, -75.460621),
    "ND": (-104.048900, -96.554507), "OH": (-84.820159, -80.518693), "OK": (-103.002455, -94.430662),
    "OR": (-124.566244, -116.463262), "PA": (-80.519891, -74.689516), "RI": (-71.886819, -71.120557),
    "SC": (-83.353910, -78.541138), "SD": (-104.057698, -96.436589), "TN": (-90.310298, -81.646900),
    "TX": (-106.645646, -93.508292), "UT": (-114.052998, -109.041058), "VT": (-73.437740, -71.464555),
    "VA": (-83.675395, -75.242266), "WA": (-124.763068, -116.915989), "WV": (-82.644739, -77.719519),
    "WI": (-92.889433, -86.763983), "WY": (-111.056888, -104.052160), "DC": (-77.119759, -76.909393),
    "PR": (-67.945404, -65.220703), "VI": (-65.013029, -64.564907), "GU": (144.618068, 144.956706),
    "AS": (-170.841600, -169.406622), "MP": (145.128345, 145.853700)
}

def validate_subscriber_file(input_csv, company_id):
    # Initialize error list
    errors = []

    # Step 1: Create or overwrite the company_id directory
    if os.path.exists(company_id):
        shutil.rmtree(company_id)
    os.makedirs(company_id)

    # Step 2: Copy input CSV to company_id with original filename
    original_filename = os.path.basename(input_csv)
    output_original_csv = os.path.join(company_id, original_filename)
    shutil.copyfile(input_csv, output_original_csv)

    # Step 3: Read the input CSV
    try:
        df = pd.read_csv(input_csv, dtype=str)
        print(f"Read CSV successfully: {len(df)} rows")
    except Exception as e:
        errors.append({"Row": "N/A", "Column": "N/A", "Error": f"Failed to read CSV: {str(e)}", "Value": "N/A"})
        save_errors_and_exit(errors, company_id, original_filename)
        return

    # Step 4: Insert OrigRowNum column
    df.insert(0, "OrigRowNum", range(1, len(df) + 1))

    # Step 5: Validate required columns (case-insensitive, already handled in VS.py)
    required_columns = EXPECTED_COLUMNS
    input_columns = df.columns.str.lower().tolist()
    missing_columns = [col for col in required_columns if col not in input_columns]
    if missing_columns:
        errors.append({
            "Row": "N/A",
            "Column": "N/A",
            "Error": f"The following required columns are missing: {', '.join(missing_columns)}",
            "Value": "N/A"
        })
        save_errors_and_exit(errors, company_id, original_filename)
        return

    # Step 6: Create cleaned DataFrame with standardized column titles
    output_columns = ["OrigRowNum"] + required_columns
    column_mapping = {col: col.lower() for col in df.columns if col.lower() in required_columns}
    column_mapping['OrigRowNum'] = 'OrigRowNum'
    try:
        cleaned_df = df[list(column_mapping.keys())].rename(columns=column_mapping)[output_columns]
    except KeyError as e:
        errors.append({
            "Row": "N/A",
            "Column": "N/A",
            "Error": f"KeyError: {e}. Available columns: {df.columns.tolist()}",
            "Value": "N/A"
        })
        save_errors_and_exit(errors, company_id, original_filename)
        return

    # Step 7: Column-based validation
    flagged_cells = {}  # Dictionary to track cells to highlight yellow: {(row, col): error_message}
    for col in cleaned_df.columns:
        if col == "OrigRowNum":
            continue
        values = cleaned_df[col].fillna("").astype(str).str.strip()

        if col == "customer":
            # Check for commas
            has_comma = values.str.contains(",", na=False)
            for idx, (val, has) in enumerate(zip(values, has_comma)):
                if has:
                    errors.append({
                        "Row": cleaned_df["OrigRowNum"][idx],
                        "Column": col,
                        "Error": "Customer ID contains a comma",
                        "Value": val
                    })
                    flagged_cells[(idx, col)] = "Customer ID contains a comma"
            # Check for duplicates
            duplicates = values[values.duplicated(keep=False)]
            for val in duplicates.unique():
                dup_indices = values[values == val].index
                for idx in dup_indices:
                    errors.append({
                        "Row": cleaned_df["OrigRowNum"][idx],
                        "Column": col,
                        "Error": "Duplicate customer ID",
                        "Value": val
                    })
                    flagged_cells[(idx, col)] = "Duplicate customer ID"

        elif col in ["lat", "lon"]:
            # Check for numeric or blank
            is_numeric = values.apply(lambda x: x == "" or pd.isna(x) or is_float(x))
            for idx, (val, valid) in enumerate(zip(values, is_numeric)):
                if val and not valid:
                    errors.append({
                        "Row": cleaned_df["OrigRowNum"][idx],
                        "Column": col,
                        "Error": f"{col.capitalize()} must be a number or blank",
                        "Value": val
                    })
                    flagged_cells[(idx, col)] = f"{col.capitalize()} must be a number or blank"
                elif val:
                    try:
                        float_val = float(val)
                        if col == "lat" and not (-90 <= float_val <= 90):
                            errors.append({
                                "Row": cleaned_df["OrigRowNum"][idx],
                                "Column": col,
                                "Error": "Latitude must be between -90 and 90",
                                "Value": val
                            })
                            flagged_cells[(idx, col)] = "Latitude must be between -90 and 90"
                        elif col == "lon":
                            state = cleaned_df["state"].iloc[idx].strip().upper() if "state" in cleaned_df else ""
                            if state in STATE_LON_RANGES:
                                lon_min, lon_max = STATE_LON_RANGES[state]
                                if state in ["GU", "MP"]:
                                    if not (lon_min <= float_val <= lon_max):
                                        errors.append({
                                            "Row": cleaned_df["OrigRowNum"][idx],
                                            "Column": col,
                                            "Error": f"Longitude for {state} must be between {lon_min} and {lon_max}",
                                            "Value": val
                                        })
                                        flagged_cells[(idx, col)] = f"Longitude for {state} must be between {lon_min} and {lon_max}"
                                else:
                                    if float_val > 0:
                                        errors.append({
                                            "Row": cleaned_df["OrigRowNum"][idx],
                                            "Column": col,
                                            "Error": f"Longitude for {state} must be negative",
                                            "Value": val
                                        })
                                        flagged_cells[(idx, col)] = f"Longitude for {state} must be negative"
                                    elif not (lon_min <= float_val <= lon_max):
                                        errors.append({
                                            "Row": cleaned_df["OrigRowNum"][idx],
                                            "Column": col,
                                            "Error": f"Longitude for {state} must be between {lon_min} and {lon_max}",
                                            "Value": val
                                        })
                                        flagged_cells[(idx, col)] = f"Longitude for {state} must be between {lon_min} and {lon_max}"
                    except ValueError:
                        pass  # Already caught by is_numeric check

        elif col == "address":
            # Check for blank or whitespace
            is_blank = values == "" 
            for idx, (val, blank) in enumerate(zip(values, is_blank)):
                if blank:
                    errors.append({
                        "Row": cleaned_df["OrigRowNum"][idx],
                        "Column": col,
                        "Error": "Blank or whitespace-only value",
                        "Value": val
                    })
                    flagged_cells[(idx, col)] = "Blank or whitespace-only value"

            # Check for PO Box
            has_po_box = values.str.contains(PO_BOX, case=False, na=False)
            for idx, (val, has) in enumerate(zip(values, has_po_box)):
                if has:
                    errors.append({
                        "Row": cleaned_df["OrigRowNum"][idx],
                        "Column": col,
                        "Error": "Address must be a physical address, PO Boxes are not allowed",
                        "Value": val
                    })
                    flagged_cells[(idx, col)] = "Address must be a physical address, PO Boxes are not allowed"

            # Check for rural routes
            has_rural_route = values.str.contains(RURAL_ROUTES, case=False, na=False)
            for idx, (val, has) in enumerate(zip(values, has_rural_route)):
                if has:
                    forbidden = re.search(FORBIDDEN_CHARS, val)
                    if forbidden:
                        errors.append({
                            "Row": cleaned_df["OrigRowNum"][idx],
                            "Column": col,
                            "Error": f"Address contains forbidden character: {forbidden.group()}",
                            "Value": val
                        })
                        flagged_cells[(idx, col)] = f"Address contains forbidden character: {forbidden.group()}"
                    continue  # Skip further checks for rural routes

            # Check for forbidden characters
            for idx, val in enumerate(values):
                forbidden = re.search(FORBIDDEN_CHARS, val)
                if forbidden:
                    errors.append({
                        "Row": cleaned_df["OrigRowNum"][idx],
                        "Column": col,
                        "Error": f"Address contains forbidden character: {forbidden.group()}",
                        "Value": val
                    })
                    flagged_cells[(idx, col)] = f"Address contains forbidden character: {forbidden.group()}"

            # Check for void/_Diamond
            has_void_diamond = values.str.contains(r"void\s+_(upload|Diamond)", case=False, na=False)
            for idx, (val, has) in enumerate(zip(values, has_void_diamond)):
                if has:
                    errors.append({
                        "Row": cleaned_df["OrigRowNum"][idx],
                        "Column": col,
                        "Error": "Contains invalid void/_Diamond code block",
                        "Value": val
                    })
                    flagged_cells[(idx, col)] = "Contains invalid void/_Diamond code block"

            # Check for specific road patterns
            has_specific_road = values.str.contains(SPECIFIC_ROAD_PATTERN, case=False, na=False)
            # Check for street endings and house numbers
            for idx, val in enumerate(values):
                if has_specific_road.iloc[idx]:
                    continue
                street_ending_match = re.search(rf"\s+(?:{STREET_ENDINGS})\.?\s*(\S.*)?$", val, re.IGNORECASE)
                if not street_ending_match:
                    errors.append({
                        "Row": cleaned_df["OrigRowNum"][idx],
                        "Column": col,
                        "Error": "Address does not match expected road or street format",
                        "Value": val
                    })
                    flagged_cells[(idx, col)] = "Address does not match expected road or street format"
                    continue
                ending = None
                for multi_word in MULTI_WORD_ENDINGS.split("|"):
                    if re.search(rf"\s+{multi_word}\.?\s*(\S.*)?$", val, re.IGNORECASE):
                        ending = re.search(rf"\s+({multi_word})\.?\s*(\S.*)?$", val, re.IGNORECASE).group(1)
                        break
                if not ending:
                    for single_word in SINGLE_WORD_ENDINGS.split("|"):
                        if re.search(rf"\s+{single_word}\.?\s*(\S.*)?$", val, re.IGNORECASE):
                            ending = re.search(rf"\s+({single_word})\.?\s*(\S.*)?$", val, re.IGNORECASE).group(1)
                            break
                if not re.search(r"^\d+", val.split(ending)[0].strip()):
                    errors.append({
                        "Row": cleaned_df["OrigRowNum"][idx],
                        "Column": col,
                        "Error": f"Address must include a house number before ending: {ending}",
                        "Value": val
                    })
                    flagged_cells[(idx, col)] = f"Address must include a house number before ending: {ending}"
                    continue
                extra = street_ending_match.group(1).strip() if street_ending_match.group(1) else ""
                if extra:
                    is_special_ending = (
                        ending.lower() in ["highway", "hwy", "county road", "county rd", "co rd", "state route", "sr",
                                           "interstate", "i-", "farm to market", "farm road", "fm", "us", "us hwy", "pvt", "private road",
                                           "county hwy", "ch", "county fm", "fm road", "fire district", "road", "rd",
                                           "route c-", "c-", "route", "rs", "ky rs", "state hwy",
                                           "az-", "ca-", "ct-", "de-", "fl-", "ga-", "id-", "il-", "in-", "k-",
                                           "me-", "md-", "ma-", "m-", "mn-", "ms-", "nh-", "nj-", "nm-", "ny-",
                                           "nc-", "oh-", "ok-", "or-", "pa-", "ri-", "sc-", "tn-", "ut-", "vt-",
                                           "va-", "wa-", "wv-", "wi-", "wy-", "sh-", "carr", "pr", "cr"] or
                        ending.lower().startswith(("route ", "county hwy ", "county fm ", "fm road ", "fire district ", "state hwy ", "ky rs ")) or
                        ending == "CR"
                    )
                    if is_special_ending:
                        if not re.match(
                            r"^(?:[0-9]+(?:\s+(?:N|S|E|W|NE|NW|SE|SW|North|South|East|West|Northeast|Northwest|Southeast|Southwest))?$|^[0-9]+$|"
                            r"[A-Za-z0-9\-]+(?:\s+(?:N|S|E|W|NE|NW|SE|SW|North|South|East|West|Northeast|Northwest|Southeast|Southwest))?|"
                            r"[A-Za-z0-9\-]+[NSEW]{1,2}|"
                            r"(?:Avenue|Ave|Av|Boulevard|Blvd|Circle|Cir|Cr|Court|Ct|Drive|Dr|Expressway|Expy|"
                            r"Highway|Hwy|Lane|Ln|Parkway|Pkwy|Place|Pl|Road|Rd|Square|Sq|Street|St|Terrace|Ter|"
                            r"Trail|Trl|Way|Wy|CR|SR|FM|US|Interstate|I-))$",
                            extra, re.IGNORECASE
                        ):
                            errors.append({
                                "Row": cleaned_df["OrigRowNum"][idx],
                                "Column": col,
                                "Error": f"Address may contain non-standard components after ending: {extra}",
                                "Value": val
                            })
                            flagged_cells[(idx, col)] = f"Address may contain non-standard components after ending: {extra}"
                    else:
                        if not re.match(r"^(?:N|S|E|W|NE|NW|SE|SW|North|South|East|West|Northeast|Northwest|Southeast|Southwest|(?:N|S|E|W)\s+(?:N|S|E|W))$", extra, re.IGNORECASE):
                            errors.append({
                                "Row": cleaned_df["OrigRowNum"][idx],
                                "Column": col,
                                "Error": f"Address may contain non-standard components after ending: {extra}",
                                "Value": val
                            })
                            flagged_cells[(idx, col)] = f"Address may contain non-standard components after ending: {extra}"

        elif col in ["city", "state", "zip", "download", "upload", "voip_lines_quantity", "business_customer", "technology"]:
            is_blank = values == ""
            for idx, (val, blank) in enumerate(zip(values, is_blank)):
                if blank and col not in ["lat", "lon"]:
                    errors.append({
                        "Row": cleaned_df["OrigRowNum"][idx],
                        "Column": col,
                        "Error": "Blank or whitespace-only value",
                        "Value": val
                    })
                    flagged_cells[(idx, col)] = "Blank or whitespace-only value"

            if col == "city":
                has_digits = values.str.contains(r"[0-9]", na=False)
                for idx, (val, has) in enumerate(zip(values, has_digits)):
                    if has:
                        errors.append({
                            "Row": cleaned_df["OrigRowNum"][idx],
                            "Column": col,
                            "Error": "City name contains digits",
                            "Value": val
                        })
                        flagged_cells[(idx, col)] = "City name contains digits"

            elif col == "state":
                invalid_states = ~values.str.upper().isin(VALID_STATES)
                for idx, (val, invalid) in enumerate(zip(values, invalid_states)):
                    if invalid:
                        errors.append({
                            "Row": cleaned_df["OrigRowNum"][idx],
                            "Column": col,
                            "Error": f"Invalid state. Must be one of {VALID_STATES}",
                            "Value": val
                        })
                        flagged_cells[(idx, col)] = f"Invalid state. Must be one of {VALID_STATES}"

            elif col == "zip":
                valid_zip = values.str.match(r"^\d{5}(-\d{4})?$")
                for idx, (val, valid) in enumerate(zip(values, valid_zip)):
                    if not valid and val:
                        errors.append({
                            "Row": cleaned_df["OrigRowNum"][idx],
                            "Column": col,
                            "Error": "Invalid ZIP code format. Must be 12345 or 12345-6789",
                            "Value": val
                        })
                        flagged_cells[(idx, col)] = "Invalid ZIP code format. Must be 12345 or 12345-6789"

            elif col in ["download", "upload"]:
                is_numeric = values.apply(lambda x: x == "" or pd.isna(x) or is_float(x))
                for idx, (val, valid) in enumerate(zip(values, is_numeric)):
                    if not valid and val:
                        errors.append({
                            "Row": cleaned_df["OrigRowNum"][idx],
                            "Column": col,
                            "Error": f"{col.capitalize()} speed must be a number",
                            "Value": val
                        })
                        flagged_cells[(idx, col)] = f"{col.capitalize()} speed must be a number"
                    elif val:
                        try:
                            speed = float(val)
                            if speed <= 0:
                                errors.append({
                                    "Row": cleaned_df["OrigRowNum"][idx],
                                    "Column": col,
                                    "Error": f"{col.capitalize()} speed must be greater than 0",
                                    "Value": val
                                })
                                flagged_cells[(idx, col)] = f"{col.capitalize()} speed must be greater than 0"
                            if speed > 3000:
                                errors.append({
                                    "Row": cleaned_df["OrigRowNum"][idx],
                                    "Column": col,
                                    "Error": f"{col.capitalize()} speed cannot exceed 3000 Mbps",
                                    "Value": val
                                })
                                flagged_cells[(idx, col)] = f"{col.capitalize()} speed cannot exceed 3000 Mbps"
                        except ValueError:
                            pass  # Already caught by is_numeric

            elif col == "voip_lines_quantity":
                is_integer = values.apply(lambda x: x == "" or pd.isna(x) or is_integer(x))
                for idx, (val, valid) in enumerate(zip(values, is_integer)):
                    if not valid and val:
                        errors.append({
                            "Row": cleaned_df["OrigRowNum"][idx],
                            "Column": col,
                            "Error": "VOIP lines quantity must be an integer",
                            "Value": val
                        })
                        flagged_cells[(idx, col)] = "VOIP lines quantity must be an integer"
                    elif val:
                        try:
                            qty = int(val)
                            if qty < 0:
                                errors.append({
                                    "Row": cleaned_df["OrigRowNum"][idx],
                                    "Column": col,
                                    "Error": "VOIP lines quantity must be non-negative",
                                    "Value": val
                                })
                                flagged_cells[(idx, col)] = "VOIP lines quantity must be non-negative"
                        except ValueError:
                            pass  # Already caught by is_integer

            elif col == "business_customer":
                valid_business = values.isin(["0", "1"])
                for idx, (val, valid) in enumerate(zip(values, valid_business)):
                    if not valid and val:
                        errors.append({
                            "Row": cleaned_df["OrigRowNum"][idx],
                            "Column": col,
                            "Error": "Business customer must be 0 or 1",
                            "Value": val
                        })
                        flagged_cells[(idx, col)] = "Business customer must be 0 or 1"

            elif col == "technology":
                valid_tech = values.str.lower().isin(VALID_TECHNOLOGIES)
                for idx, (val, valid) in enumerate(zip(values, valid_tech)):
                    if not valid and val:
                        errors.append({
                            "Row": cleaned_df["OrigRowNum"][idx],
                            "Column": col,
                            "Error": f"Invalid technology. Must be one of {VALID_TECHNOLOGIES}",
                            "Value": val
                        })
                        flagged_cells[(idx, col)] = f"Invalid technology. Must be one of {VALID_TECHNOLOGIES}"

    # Step 8: Save cleaned DataFrame
    base_filename = os.path.splitext(original_filename)[0]
    output_cleantitles_csv = os.path.join(company_id, f"{base_filename}_Mod_1.csv")
    try:
        cleaned_df.to_csv(output_cleantitles_csv, index=False)
        if os.path.isfile(output_cleantitles_csv):
            print(f"Successfully saved: {output_cleantitles_csv}")
        else:
            errors.append({
                "Row": "N/A",
                "Column": "N/A",
                "Error": f"Failed to save {output_cleantitles_csv}. File does not exist.",
                "Value": "N/A"
            })
            save_errors_and_exit(errors, company_id, original_filename)
            return
    except Exception as e:
        errors.append({
            "Row": "N/A",
            "Column": "N/A",
            "Error": f"Error saving {output_cleantitles_csv}: {str(e)}",
            "Value": "N/A"
        })
        save_errors_and_exit(errors, company_id, original_filename)
        return

    # Step 9: Save errors to CSV
    errors_csv_path = os.path.join(company_id, f"{base_filename}_Errors.csv")
    df_errors = pd.DataFrame(errors)
    if not df_errors.empty:
        df_errors = df_errors.sort_values(by="Error")
        df_errors.to_csv(errors_csv_path, index=False)
    else:
        pd.DataFrame(columns=["Row", "Column", "Error", "Value"]).to_csv(errors_csv_path, index=False)
    print(f"Errors CSV saved: {errors_csv_path}")

    # Step 10: Save Corrected_Subscribers.csv with cell coloring
    corrected_csv_path = os.path.join(company_id, f"{base_filename}_Corrected_Subscribers.csv")
    try:
        # Save DataFrame to CSV first
        cleaned_df.to_csv(corrected_csv_path, index=False)
        # Load with openpyxl to apply cell colors
        wb = openpyxl.load_workbook(corrected_csv_path)
        ws = wb.active
        # Define fill colors
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        # Map column names to Excel columns (A=1, B=2, etc.)
        col_map = {col: idx + 1 for idx, col in enumerate(cleaned_df.columns)}
        # Apply yellow fill to flagged cells
        for (row_idx, col_name), _ in flagged_cells.items():
            excel_col = openpyxl.utils.get_column_letter(col_map[col_name])
            excel_row = row_idx + 2  # +1 for header, +1 for 1-based indexing
            ws[f"{excel_col}{excel_row}"].fill = yellow_fill
        wb.save(corrected_csv_path)
        if os.path.isfile(corrected_csv_path):
            print(f"Successfully saved: {corrected_csv_path}")
        else:
            errors.append({
                "Row": "N/A",
                "Column": "N/A",
                "Error": f"Failed to save {corrected_csv_path}. File does not exist.",
                "Value": "N/A"
            })
            save_errors_and_exit(errors, company_id, original_filename)
            return
    except Exception as e:
        errors.append({
            "Row": "N/A",
            "Column": "N/A",
            "Error": f"Error saving {corrected_csv_path}: {str(e)}",
            "Value": "N/A"
        })
        save_errors_and_exit(errors, company_id, original_filename)
        return

    # Step 11: Print summary
    print(f"Processing complete. Files saved in {company_id}/:")
    print(f"- {original_filename} (original copy)")
    print(f"- {base_filename}_Mod_1.csv (cleaned column titles with OrigRowNum)")
    print(f"- {base_filename}_Errors.csv (validation errors)")
    print(f"- {base_filename}_Corrected_Subscribers.csv (data with flagged cells in yellow)")
    print(f"Total rows: {len(cleaned_df)}, Flagged cells: {len(flagged_cells)}")

def save_errors_and_exit(errors, company_id, original_filename):
    base_filename = os.path.splitext(original_filename)[0]
    errors_csv_path = os.path.join(company_id, f"{base_filename}_Errors.csv")
    pd.DataFrame(errors).to_csv(errors_csv_path, index=False)
    print(f"Errors CSV saved: {errors_csv_path}")
    print("Processing terminated due to errors.")
    sys.exit(1)

def is_float(value):
    try:
        float(value)
        return True
    except ValueError:
        return False

def is_integer(value):
    try:
        int(value)
        return float(value).is_integer()
    except ValueError:
        return False

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python3 VS.py <input_csv> <company_id>")
        sys.exit(1)

    input_csv = sys.argv[1]
    company_id = sys.argv[2]

    if not os.path.isfile(input_csv):
        print(f"Error: Input file '{input_csv}' does not exist.")
        sys.exit(1)

    validate_subscriber_file(input_csv, company_id)